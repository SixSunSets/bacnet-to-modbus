import openpyxl
import json
from collections import OrderedDict

# Cargar el archivo Excel
arch_excel = r"C:/Users/User/Desktop/bacnet-to-modbus/Lista_de_Puntos_Daikin.xlsx"
libro = openpyxl.load_workbook(arch_excel)
nombres_hojas = libro.sheetnames[0:1]  # Selecciona la primera hoja
print(nombres_hojas)

def excel_a_json():
    equipos_ac = OrderedDict()
    
    for hoja in nombres_hojas:
        tabla = libro[hoja]
        equipos_ac[hoja] = OrderedDict()  
        
        for fila in tabla.iter_rows(values_only=True):
            if fila[0] == "SI" and isinstance(fila[1], int): 
                id = fila[1]
                if id not in equipos_ac[hoja]:
                    equipos_ac[hoja][id] = []  # Crear una lista para el ID si no existe
                equipos_ac[hoja][id].append(fila)  # Agregar la fila a la lista del ID

    # Guardar el JSON
    with open(r"C:/Users/User/Desktop/bacnet-to-modbus/Lista_de_Puntos_Daikin.json", "w") as archivo_json:
        json.dump(equipos_ac, archivo_json, indent=4) 

excel_a_json()